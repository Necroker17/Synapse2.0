"""
Genera synapse_business_plan.xlsx a partir de los documentos de business_plan/.
Los .md siguen siendo la fuente de verdad; re-ejecutar este script tras cualquier
cambio en pricing_strategy.md, ambassador_program.md, ambassador_scenarios.md,
financial_model.md o profit_split.md.

Uso: python3 build_business_plan_xlsx.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PENDING_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
CURRENCY = '"$"#,##0.00'
PERCENT = "0%"


def style_header_row(ws, row, first_col, last_col):
    for col in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_pricing(wb):
    ws = wb.active
    ws.title = "Pricing"
    ws["A1"] = "Planes de Membresía — Synapse AI Scanner"
    ws["A1"].font = Font(bold=True, size=13)

    headers = ["Plan", "Duración (meses)", "Precio Lanzamiento (USD)", "Precio Oficial (USD)", "Entregables clave"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, 1, len(headers))

    rows = [
        ("STANDARD", 3, 197, 249,
         "Indicador (3m); comunidad privada WhatsApp (señales); clases en vivo 1-2x/semana; todos los módulos de Skool"),
        ("PRO", 6, 299, 397,
         "Todo lo de STANDARD + indicador (6m) + señales VIP de altos ratios + soporte prioritario"),
        ("PREMIUM", 12, 699, 799,
         "Todo lo de PRO + indicador (12m) + Grupo Elite (mentoría 1 a 1)"),
    ]
    for i, (plan, dur, launch, official, deliverables) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=plan).font = BOLD
        ws.cell(row=i, column=2, value=dur)
        c = ws.cell(row=i, column=3, value=launch)
        c.number_format = CURRENCY
        c = ws.cell(row=i, column=4, value=official)
        c.number_format = CURRENCY
        ws.cell(row=i, column=5, value=deliverables).alignment = Alignment(wrap_text=True, vertical="top")

    ws["A8"] = "Nota: corte del precio de lanzamiento es por fecha fija (aún sin definir); política de renovación (precio oficial vs. de adquisición) pendiente. Ver pricing_strategy.md sección 3."
    ws["A8"].font = Font(italic=True, size=9)
    ws.merge_cells("A8:E8")
    ws["A8"].alignment = Alignment(wrap_text=True)

    autosize(ws, [14, 16, 20, 18, 55])
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 30
    return ws


def build_embudo(wb):
    ws = wb.create_sheet("Embudo Trial")
    ws["A1"] = "Embudo de Adquisición — Free Trial"
    ws["A1"].font = Font(bold=True, size=13)

    headers = ["Elemento", "Detalle"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, 1, len(headers))

    rows = [
        ("Duración", "15 días de acceso gratuito al indicador"),
        ("Incentivo económico", "Bono de $30 USD otorgado por el BROKER al registrarse (no es costo de Synapse)"),
        ("Educación", "Acceso al Módulo 1 de Skool (Trading básico)"),
        ("Comunidad", "Grupo abierto de WhatsApp: clases gratuitas + testimonios/casos de éxito diarios"),
    ]
    for i, (elem, detalle) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=elem).font = BOLD
        ws.cell(row=i, column=2, value=detalle).alignment = Alignment(wrap_text=True)

    autosize(ws, [22, 70])
    return ws


def build_comisiones(wb, price_cell_launch, price_cell_official):
    ws = wb.create_sheet("Comisiones Embajadores")
    ws["A1"] = "Estructura de Comisión de Embajadores"
    ws["A1"].font = Font(bold=True, size=13)

    headers = ["Evento", "Tasa de comisión"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, 1, len(headers))

    ws.cell(row=4, column=1, value="Primera venta de un cliente referido").font = BOLD
    c = ws.cell(row=4, column=2, value=0.5)
    c.number_format = PERCENT
    ws.cell(row=5, column=1, value="Segunda re-compra en adelante (re-consumo)").font = BOLD
    c = ws.cell(row=5, column=2, value=0.25)
    c.number_format = PERCENT

    ws["A7"] = "Ejemplo de comisión en USD por plan (vinculado a la pestaña Pricing)"
    ws["A7"].font = BOLD
    headers2 = ["Plan", "Precio Lanzamiento", "Precio Oficial", "Comisión 1ra venta (lanz.)", "Comisión 1ra venta (of.)", "Comisión re-consumo (lanz.)", "Comisión re-consumo (of.)"]
    for col, h in enumerate(headers2, start=1):
        ws.cell(row=8, column=col, value=h)
    style_header_row(ws, 8, 1, len(headers2))

    plans = ["STANDARD", "PRO", "PREMIUM"]
    for i, plan in enumerate(plans, start=9):
        pricing_row = i - 9 + 4  # Pricing sheet rows start at 4
        ws.cell(row=i, column=1, value=plan).font = BOLD
        ws.cell(row=i, column=2, value=f"='Pricing'!C{pricing_row}").number_format = CURRENCY
        ws.cell(row=i, column=3, value=f"='Pricing'!D{pricing_row}").number_format = CURRENCY
        ws.cell(row=i, column=4, value=f"=B{i}*$B$4").number_format = CURRENCY
        ws.cell(row=i, column=5, value=f"=C{i}*$B$4").number_format = CURRENCY
        ws.cell(row=i, column=6, value=f"=B{i}*$B$5").number_format = CURRENCY
        ws.cell(row=i, column=7, value=f"=C{i}*$B$5").number_format = CURRENCY

    autosize(ws, [16, 18, 16, 22, 20, 24, 22])
    return ws


def build_escenarios(wb):
    ws = wb.create_sheet("Escenarios Embajadores")
    ws["A1"] = "Matriz de Escenarios — Ingreso Anual de UN Embajador"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "Fórmula: comisión_anual = n × P × [6 + 2.25r + 1.5r² + 0.75r³]  (n=nuevos/mes, P=precio plan, r=retención trimestral)"
    ws["A2"].font = Font(italic=True, size=9)
    ws.merge_cells("A2:F2")

    def price_table(start_row, title, price_formula):
        ws.cell(row=start_row, column=1, value=title).font = BOLD
        ws.cell(row=start_row, column=6, value="Precio usado:")
        ws.cell(row=start_row, column=6).font = Font(italic=True, size=9)
        price_cell = f"G{start_row}"
        ws[price_cell] = price_formula
        ws[price_cell].number_format = CURRENCY

        header_row = start_row + 1
        ws.cell(row=header_row, column=1, value="Volumen/mes (n)")
        r_values = [0.20, 0.40, 0.60]
        r_labels = ["Conservador (20%)", "Base (40%)", "Optimista (60%)"]
        for col, (r, label) in enumerate(zip(r_values, r_labels), start=2):
            ws.cell(row=header_row, column=col, value=label)
            ws.cell(row=header_row + 1, column=col, value=r).number_format = PERCENT
        style_header_row(ws, header_row, 1, 4)
        ws.cell(row=header_row + 1, column=1, value="(r usado por columna →)").font = Font(italic=True, size=8)

        n_values = [3, 5, 10, 15, 20]
        n_labels = ["Principiante", "Promedio", "Consistente", "Alto rendimiento", "Élite"]
        data_start = header_row + 2
        for i, (n, label) in enumerate(zip(n_values, n_labels)):
            row = data_start + i
            ws.cell(row=row, column=1, value=f"{label} ({n}/mes)")
            n_cell_col = get_column_letter(6)  # store literal n in helper col G below the price, use col H instead
        # Store n helper values in column H (hidden-ish) so formulas can reference a numeric cell
        helper_col = 8  # column H
        for i, n in enumerate(n_values):
            ws.cell(row=data_start + i, column=helper_col, value=n)
        for i in range(len(n_values)):
            row = data_start + i
            for col in range(2, 5):
                r_row = header_row + 1
                r_ref = f"{get_column_letter(col)}${r_row}"
                n_ref = f"$H{row}"
                formula = f"={n_ref}*${price_cell}*(6+2.25*{r_ref}+1.5*{r_ref}^2+0.75*{r_ref}^3)"
                c = ws.cell(row=row, column=col, value=formula)
                c.number_format = CURRENCY
        return data_start + len(n_values)

    next_row = price_table(4, "Precio Oficial STANDARD", "='Pricing'!D4")
    price_table(next_row + 2, "Precio Lanzamiento STANDARD", "='Pricing'!C4")

    autosize(ws, [26, 20, 20, 20, 2, 14, 12, 8])
    ws.column_dimensions["H"].hidden = False  # keep visible; it's the n helper column, useful for transparency
    return ws


def build_financiero(wb):
    ws = wb.create_sheet("Modelo Financiero")
    ws["A1"] = "Modelo Financiero — Costos y Margen"
    ws["A1"].font = Font(bold=True, size=13)

    ws["A3"] = "Costos fijos mensuales del negocio (confirmados 2026-07-07)"
    ws["A3"].font = BOLD
    headers = ["Herramienta", "Costo mensual (USD)"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, 1, 2)
    tools = [("TradingView Premium", 85), ("Make (automatizaciones)", 10), ("Manychat", 25), ("Skool", 10), ("Claude", 20)]
    for i, (tool, cost) in enumerate(tools, start=5):
        ws.cell(row=i, column=1, value=tool)
        ws.cell(row=i, column=2, value=cost).number_format = CURRENCY
    total_row = 5 + len(tools)
    ws.cell(row=total_row, column=1, value="Total costo_fijo_mensual_negocio").font = BOLD
    c = ws.cell(row=total_row, column=2, value=f"=SUM(B5:B{total_row - 1})")
    c.number_format = CURRENCY
    c.font = BOLD

    pending_start = total_row + 2
    ws.cell(row=pending_start, column=1, value="Costos aún sin confirmar (no inventar — dejar en 0 hasta tener el dato real)").font = BOLD
    headers2 = ["Variable", "Qué representa", "Valor"]
    for col, h in enumerate(headers2, start=1):
        ws.cell(row=pending_start + 1, column=col, value=h)
    style_header_row(ws, pending_start + 1, 1, 3)
    pending_vars = [
        ("fee_pago", "% de comisión de la pasarela de pago (Stripe, etc.)", 0),
        ("costo_soporte_mensual", "Costo de soporte/comunidad por cliente activo al mes", 0),
        ("costo_herramientas_mensual_por_cliente", "Prorrateo por cliente de alguna herramienta que escale con clientes", 0),
    ]
    pending_price_row = pending_start + 2
    for i, (var, desc, val) in enumerate(pending_vars):
        row = pending_price_row + i
        ws.cell(row=row, column=1, value=var)
        ws.cell(row=row, column=2, value=desc).alignment = Alignment(wrap_text=True)
        c = ws.cell(row=row, column=3, value=val)
        c.fill = PENDING_FILL
    fee_pago_cell = f"C{pending_price_row}"
    costo_soporte_cell = f"C{pending_price_row + 1}"
    costo_herr_cell = f"C{pending_price_row + 2}"

    margen_start = pending_price_row + len(pending_vars) + 2
    ws.cell(row=margen_start, column=1, value="Margen por venta individual (referido por embajador, primera venta, plan STANDARD, precio oficial)").font = BOLD
    labels_formulas = [
        ("Precio del plan", "='Pricing'!D4"),
        ("Ingreso neto de pago", f"=B{margen_start+1}*(1-{fee_pago_cell})"),
        ("Comisión embajador (1ra venta, 50%)", f"=B{margen_start+1}*0.5"),
        ("Costo operativo del plan (soporte+herramientas × 3 meses)", f"=({costo_soporte_cell}+{costo_herr_cell})*3"),
        ("Margen neto", f"=B{margen_start+2}-B{margen_start+3}-B{margen_start+4}"),
    ]
    for i, (label, formula) in enumerate(labels_formulas):
        row = margen_start + 1 + i
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=formula)
        c.number_format = CURRENCY

    autosize(ws, [45, 45, 14])
    return ws, fee_pago_cell


def build_reparto(wb):
    ws = wb.create_sheet("Reparto Socios")
    ws["A1"] = "Reparto de Ganancia Libre"
    ws["A1"].font = Font(bold=True, size=13)

    ws["A3"] = "Ganancia libre anual (input manual — pendiente de cifra real confirmada; costos de fee de pago y soporte de Modelo Financiero aún no están incluidos en ningún número real)"
    ws["A3"].font = Font(italic=True, size=9)
    ws.merge_cells("A3:D3")
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws["A4"] = "Ganancia libre anual (USD)"
    ws["A4"].font = BOLD
    ws["B4"] = 0
    ws["B4"].fill = PENDING_FILL
    ws["B4"].number_format = CURRENCY

    headers = ["Destino", "% de la ganancia libre", "Monto (USD)"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=6, column=col, value=h)
    style_header_row(ws, 6, 1, 3)
    rows = [
        ("Gustavo (estratega digital)", 0.25),
        ("Reserva (impuestos/imprevistos/stack de plataformas)", 0.15),
        ("Fundador", 0.30),
        ("Meli (rol comercial)", 0.30),
    ]
    for i, (dest, pct) in enumerate(rows, start=7):
        ws.cell(row=i, column=1, value=dest)
        c = ws.cell(row=i, column=2, value=pct)
        c.number_format = PERCENT
        c = ws.cell(row=i, column=3, value=f"=$B$4*B{i}")
        c.number_format = CURRENCY
    total_row = 7 + len(rows)
    ws.cell(row=total_row, column=1, value="Total").font = BOLD
    c = ws.cell(row=total_row, column=2, value=f"=SUM(B7:B{total_row-1})")
    c.number_format = PERCENT
    c.font = BOLD
    c = ws.cell(row=total_row, column=3, value=f"=SUM(C7:C{total_row-1})")
    c.number_format = CURRENCY
    c.font = BOLD

    ex_start = total_row + 3
    ws.cell(row=ex_start, column=1, value="Ejemplo ilustrativo ya calculado en la conversación (escenario: 10 embajadores × 10 nuevos/mes, 40% retención trimestral, STANDARD) — NO es una cifra real medida").font = Font(italic=True, size=9)
    ws.merge_cells(f"A{ex_start}:D{ex_start}")
    ws.cell(row=ex_start, column=1).alignment = Alignment(wrap_text=True)
    headers3 = ["Escenario", "Ganancia libre anual estimada"]
    for col, h in enumerate(headers3, start=1):
        ws.cell(row=ex_start + 1, column=col, value=h)
    style_header_row(ws, ex_start + 1, 1, 2)
    examples = [("Precio de lanzamiento", 186610.8), ("Precio oficial", 238143.6)]
    for i, (label, val) in enumerate(examples, start=ex_start + 2):
        ws.cell(row=i, column=1, value=label)
        c = ws.cell(row=i, column=2, value=val)
        c.number_format = CURRENCY

    autosize(ws, [45, 22, 18])
    return ws


def main():
    wb = Workbook()
    build_pricing(wb)
    build_embudo(wb)
    build_comisiones(wb, "Pricing!C4", "Pricing!D4")
    build_escenarios(wb)
    build_financiero(wb)
    build_reparto(wb)

    out_path = "synapse_business_plan.xlsx"
    wb.save(out_path)
    print(f"OK: generado {out_path}")


if __name__ == "__main__":
    main()
